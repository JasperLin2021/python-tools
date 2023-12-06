import os
import re
from datetime import datetime

import pandas as pd

import openpyxl
from openpyxl.formatting.rule import IconSetRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from utils import deleteRow, get_ad_sku_dict
from openpyxl import load_workbook, Workbook


def createDirectory(folder_name):
    # 获取当前工作目录
    current_directory = os.getcwd()

    folder_path = os.path.join(current_directory, folder_name)
    os.makedirs(folder_path)


def arrangeCostHeadProcess_Site():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出2"

    # 创建一个新的Excel文件来保存数据透视表
    output_filename1 = "Arrange_成本头程_站点1.xlsx"
    output_filename2 = "Arrange_成本头程_站点2.xlsx"
    new_file1 = os.path.join(current_directory, folder_name, output_filename1)
    new_file2 = os.path.join(current_directory, folder_name, output_filename2)

    target_workbook1 = openpyxl.Workbook()
    target_workbook2 = openpyxl.Workbook()
    target_sheet1 = target_workbook1.active
    target_sheet2 = target_workbook2.active

    files = [file for file in os.listdir() if file.startswith('成本头程')]
    for file in files:
        # 读取原始的Excel文件
        wb = load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=1, values_only=True):
            data_a = row[0]
            data_k = row[10]
            data_l = row[11]
            data_m = row[12]
            data_o = row[14]
            data_p = row[15]

            # 复制到目标文件，并在D列和F列之间插入空白列
            target_sheet1.append([data_a, data_k, data_l, data_m])
            target_sheet2.append([data_o, data_p])

        # 将B1单元格的值设置为"SKU"
        target_sheet2['B1'] = 'SKU'

    # 保存目标文件

    target_workbook1.save(new_file1)
    target_workbook2.save(new_file2)


def vlookupCostHeadProcess_Site():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出2"

    # 创建一个新的Excel文件来保存数据透视表
    output_filename1 = "Arrange_成本头程_站点1.xlsx"
    output_filename2 = "Arrange_成本头程_站点2.xlsx"
    output_filename3 = "Arrange_end_成本头程_站点.xlsx"
    file1 = os.path.join(current_directory, folder_name, output_filename1)
    file2 = os.path.join(current_directory, folder_name, output_filename2)
    new_file = os.path.join(current_directory, folder_name, output_filename3)

    # 读取abc.xlsx文件中的数据
    df_file1 = pd.read_excel(file1)

    # 读取efg.xlsx文件中的数据
    df_file2 = pd.read_excel(file2)

    # 进行VLOOKUP操作
    merged_df = pd.merge(df_file1, df_file2, how='left', on='SKU')

    merged_df.to_excel(new_file, sheet_name='成本头程_站点', index=False)


def arrangeFBA():
    # 打开Excel文件
    workbook = openpyxl.load_workbook('模板-亚马逊库存分析.xlsx')
    # 获取Sheet1工作表
    sheet = workbook['Sheet1']
    # 获取AO2单元格对象
    cell = sheet['AO2']
    first_day_of_current_month = cell.value.date()

    # 关闭Excel文件
    workbook.close()

    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name1 = "输出"
    folder_name2 = "输出2"

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "Arrange_fba库存.xlsx"
    new_file = os.path.join(current_directory, folder_name2, output_filename)

    #
    CostHeadProcess_Site_filename = "Arrange_end_成本头程_站点.xlsx"
    Currency_exchange_rate_filename = "品牌广告明细sku.xlsx"
    CostHeadProcess_Site_file = os.path.join(current_directory, folder_name2, CostHeadProcess_Site_filename)
    Cost_dict = get_ad_sku_dict(CostHeadProcess_Site_file, "成本头程_站点", 4, 1)
    HeadProcess_dict = get_ad_sku_dict(CostHeadProcess_Site_file, "成本头程_站点", 4, 2)
    Site_dict = get_ad_sku_dict(CostHeadProcess_Site_file, "成本头程_站点", 4, 3)
    Currency_exchange_rate_dict = get_ad_sku_dict(Currency_exchange_rate_filename, "币种汇率", 0, 1)

    #
    Startsellingdate_filename = "亚马逊产品开始出售日期.xlsx"
    Startsellingdate_dict = get_ad_sku_dict(Startsellingdate_filename, "Sheet1", 1, 3)
    Salesperson_dict = get_ad_sku_dict(Startsellingdate_filename, "Sheet1", 1, 4)

    Summary_filename = "汇总.xlsx"
    Summary_file = os.path.join(current_directory, folder_name1, Summary_filename)

    Seven_day_sales_dict = get_ad_sku_dict(Summary_file, "汇总", 73, 74)
    fifteen_day_sales_dict = get_ad_sku_dict(Summary_file, "汇总", 78, 79)
    thirty_day_sales_dict = get_ad_sku_dict(Summary_file, "汇总", 83, 84)

    last_week_seven_day_sales_dict = get_ad_sku_dict(Summary_file, "汇总", 50, 51)
    last_week_inventory_indicators_dict = get_ad_sku_dict(Summary_file, "汇总", 30, 31)

    Seven_day_advertisement_cost = get_ad_sku_dict(Summary_file, "汇总", 55, 57)
    Fifteen_day_advertisement_cost = get_ad_sku_dict(Summary_file, "汇总", 61, 63)
    Thirty_day_advertisement_cost = get_ad_sku_dict(Summary_file, "汇总", 67, 69)

    Seven_day_advertisement_total_sales = get_ad_sku_dict(Summary_file, "汇总", 55, 56)
    Fifteen_day_advertisement_total_sales = get_ad_sku_dict(Summary_file, "汇总", 61, 62)
    Thirty_day_advertisement_total_sales = get_ad_sku_dict(Summary_file, "汇总", 67, 68)

    last_week_seven_day_ACOS_dict = get_ad_sku_dict(Summary_file, "汇总", 40, 41)

    Last_week_end_value = get_ad_sku_dict(Summary_file, "汇总", 6, 7)
    Last_2week_end_value = get_ad_sku_dict(Summary_file, "汇总", 1, 2)

    Fifteen_day_total_sales = get_ad_sku_dict(Summary_file, "汇总", 19, 21)

    target_workbook = openpyxl.Workbook()
    target_sheet = target_workbook.active

    files = [file for file in os.listdir() if file.startswith('fba')]
    for file in files:
        # 读取原始的Excel文件
        wb = load_workbook(file)
        sheet = wb.active

        pattern = r'Y(\d+)'
        for row in sheet.iter_rows(min_row=2, values_only=True):
            match = re.search(pattern, row[1])
            if match:
                data_b = match.group(1) + "店"
            else:
                data_b = row[1]

            data_d = row[3]  # 获取D列数据
            data_m = row[12]  # 获取M列数据
            data_f = row[5]  # 获取F列数据
            data_z = row[25]  # 获取Z列数据
            data_aa = row[26]  # 获取AA列数据
            data_ac = row[28]  # 获取AC列数据
            data_ad = row[29]  # 获取AD列数据
            data_ak = row[36]  # 获取AK列数据

            operate = Salesperson_dict.get(data_d)[0] if Salesperson_dict.get(data_d) != None else '-'

            start_selling_date = Startsellingdate_dict.get(data_d)[0].date().strftime(
                "%Y/%m/%d") if Startsellingdate_dict.get(data_d) is not None and Startsellingdate_dict.get(data_d) != [
                None] else '-'
            start_selling_date_raw = Startsellingdate_dict.get(data_d)[0].date() if Startsellingdate_dict.get(
                data_d) is not None and Startsellingdate_dict.get(data_d) != [None] else '-'
            available_stock = data_z + data_aa + data_ac + data_ad + data_ak
            average_cost_raw = Cost_dict.get(data_d, [0])[0]
            average_cost = round(average_cost_raw, 2)
            end_value_raw = available_stock * average_cost_raw
            end_value = round(end_value_raw, 2)
            Seven_day_sales = Seven_day_sales_dict.get(data_d, [0])[0]
            fifteen_day_sales = fifteen_day_sales_dict.get(data_d, [0])[0]
            thirty_day_sales = thirty_day_sales_dict.get(data_d, [0])[0]
            last_week_seven_day_sales = last_week_seven_day_sales_dict.get(data_d, [0])[0]

            sales_difference_between_two_weeks_and_seven_days = Seven_day_sales - last_week_seven_day_sales

            available_days_for_sale_t30 = round(float(available_stock) / ((float(Seven_day_sales) / 7 + float(
                fifteen_day_sales) / 15 + float(thirty_day_sales) / 30) / 3), 0) if (float(Seven_day_sales) / 7 + float(
                fifteen_day_sales) / 15 + float(thirty_day_sales) / 30) != 0 else '-'
            sell_out_indicator_within_30_days = round(float(available_stock) / ((float(Seven_day_sales) / 7 + float(
                fifteen_day_sales) / 15 + float(thirty_day_sales) / 30) / 3) / 30,
                                                      2) if available_days_for_sale_t30 != '-' else '-'
            last_week_inventory_indicators = round(float(last_week_inventory_indicators_dict.get(data_d, ['-'])[0]),
                                                   2) if last_week_inventory_indicators_dict.get(data_d, ['-'])[
                                                             0] != '-' else '-'
            comparison_between_this_week_and_last_week = round((float(available_stock) / ((float(
                Seven_day_sales) / 7 + float(fifteen_day_sales) / 15 + float(thirty_day_sales) / 30) / 3) / 30) \
                                                               - \
                                                               (float(last_week_inventory_indicators_dict.get(data_d,
                                                                                                              ['-'])[
                                                                          0])), 2) \
                if sell_out_indicator_within_30_days != '-' and last_week_inventory_indicators != '-' else '-'

            diff_days = (
                        start_selling_date_raw - first_day_of_current_month).days if start_selling_date_raw != '-' else '-'

            # if diff_days == '-':
            #     inventory_alarm_in_the_past_7_days = "-"
            if diff_days != '-' and abs(diff_days) >= 0 and abs(diff_days) <= 30:
                inventory_alarm_in_the_past_7_days = "新品"
            elif available_stock > 0 and Seven_day_sales == 0:
                inventory_alarm_in_the_past_7_days = "有库存无销量"
            elif available_stock == 0 and Seven_day_sales == 0:
                inventory_alarm_in_the_past_7_days = "无库存无销量"
            elif available_stock == 0 and Seven_day_sales:
                inventory_alarm_in_the_past_7_days = "无库存有销量"
            elif (available_stock / (Seven_day_sales / 7)) / 30 <= 0.5:
                inventory_alarm_in_the_past_7_days = "库存过低"
            elif (available_stock / (Seven_day_sales / 7)) / 30 <= 1.5:
                inventory_alarm_in_the_past_7_days = "正常"
            else:
                inventory_alarm_in_the_past_7_days = "库存过高"

            # if diff_days == '-':
            #     inventory_alarm_over_45_days = "-"
            if diff_days != '-' and abs(diff_days) >= 0 and abs(diff_days) <= 30:
                inventory_alarm_over_45_days = "新品"
            elif sell_out_indicator_within_30_days == '-' and available_stock > 0:
                inventory_alarm_over_45_days = "有库存无销量"
            elif sell_out_indicator_within_30_days == '-' and available_stock == 0:
                inventory_alarm_over_45_days = "无库存无销量"
            elif sell_out_indicator_within_30_days <= 0.5 and available_stock == 0:
                inventory_alarm_over_45_days = "无库存有销量"
            elif sell_out_indicator_within_30_days <= 0.5:
                inventory_alarm_over_45_days = "库存过低"
            elif sell_out_indicator_within_30_days <= 1.5:
                inventory_alarm_over_45_days = "正常"
            else:
                inventory_alarm_over_45_days = "库存过高"

            seven_day_advertisement_cost_raw = sum(
                Seven_day_advertisement_cost.get(data_d)) if Seven_day_advertisement_cost.get(data_d) != None else 0
            seven_day_advertisement_cost = round(seven_day_advertisement_cost_raw, 2)

            fifteen_day_advertisement_cost_raw = sum(
                Fifteen_day_advertisement_cost.get(data_d)) if Fifteen_day_advertisement_cost.get(data_d) != None else 0
            fifteen_day_advertisement_cost = round(fifteen_day_advertisement_cost_raw, 2)

            thirty_day_advertisement_cost_raw = sum(
                Thirty_day_advertisement_cost.get(data_d)) if Thirty_day_advertisement_cost.get(data_d) != None else 0
            thirty_day_advertisement_cost = round(thirty_day_advertisement_cost_raw, 2)

            seven_day_total_sum_sales = sum(
                Seven_day_advertisement_total_sales.get(data_d)) if Seven_day_advertisement_total_sales.get(
                data_d) != None else 0
            seven_day_acos_raw = seven_day_advertisement_cost_raw / seven_day_total_sum_sales if seven_day_total_sum_sales != 0 and seven_day_advertisement_cost_raw != 0 else '-'
            seven_day_acos = round(seven_day_acos_raw,
                                   4) if seven_day_total_sum_sales != 0 and seven_day_advertisement_cost_raw != 0 else '-'

            last_week_seven_day_ACOS_raw = last_week_seven_day_ACOS_dict.get(data_d, ['-'])[0]
            last_week_seven_day_ACOS = round(last_week_seven_day_ACOS_raw,
                                             4) if last_week_seven_day_ACOS_raw != '-' else '-'

            two_week_ACOS_difference = round((seven_day_acos_raw - last_week_seven_day_ACOS_raw),
                                             4) if seven_day_acos_raw != '-' and last_week_seven_day_ACOS_raw != '-' else '-'

            fifteen_day_total_sum_sales = sum(
                Fifteen_day_advertisement_total_sales.get(data_d)) if Fifteen_day_advertisement_total_sales.get(
                data_d) != None else 0
            fifteen_day_acos_raw = fifteen_day_advertisement_cost_raw / fifteen_day_total_sum_sales if fifteen_day_total_sum_sales != 0 and fifteen_day_advertisement_cost_raw != 0 else '-'
            fifteen_day_acos = round(fifteen_day_acos_raw, 4) if fifteen_day_acos_raw != "-" else '-'

            thirty_day_total_sum_sales = sum(
                Thirty_day_advertisement_total_sales.get(data_d)) if Thirty_day_advertisement_total_sales.get(
                data_d) != None else 0
            thirty_day_acos_raw = thirty_day_advertisement_cost_raw / thirty_day_total_sum_sales if thirty_day_total_sum_sales != 0 and thirty_day_advertisement_cost_raw != 0 else '-'
            thirty_day_acos = round(thirty_day_acos_raw, 4) if thirty_day_acos_raw != '-' else '-'

            seven_day_sales_cost_raw = average_cost_raw * Seven_day_sales
            seven_day_sales_cost = round(seven_day_sales_cost_raw, 2)
            fifteen_day_sales_cost_raw = average_cost_raw * fifteen_day_sales
            fifteen_day_sales_cost = round(fifteen_day_sales_cost_raw, 2)

            last_week_end_value_raw = sum(Last_week_end_value.get(data_d)) if Last_week_end_value.get(
                data_d) != None else 0
            seven_day_inventory_turnover_rate = round(
                (seven_day_sales_cost_raw / ((last_week_end_value_raw + end_value_raw) / 2)),
                1) if last_week_end_value_raw + end_value_raw != 0 else '-'

            last_2week_end_value_raw = sum(Last_2week_end_value.get(data_d)) if Last_2week_end_value.get(
                data_d) != None else 0
            fifteen_day_inventory_turnover_rate_raw = (fifteen_day_sales_cost_raw / ((
                                                                                                 last_2week_end_value_raw + end_value_raw) / 2)) if last_2week_end_value_raw + end_value_raw != 0 else '-'
            fifteen_day_inventory_turnover_rate = round(fifteen_day_inventory_turnover_rate_raw,
                                                        1) if fifteen_day_inventory_turnover_rate_raw != "-" else '-'

            site = Site_dict.get(data_d, ['-'])[0]

            gross_margin_raw = (sum(Fifteen_day_total_sales.get(data_d)) * 0.85 - (
                        average_cost_raw + HeadProcess_dict.get(data_d, [0])[0]) * fifteen_day_sales -
                                Currency_exchange_rate_dict.get(site, [0])[0] *
                                (fifteen_day_advertisement_cost_raw + 3.99 * fifteen_day_sales)) / sum(
                Fifteen_day_total_sales.get(data_d)) if Fifteen_day_total_sales.get(data_d) != None and sum(
                Fifteen_day_total_sales.get(data_d)) != 0 else '-'

            gross_margin = round(gross_margin_raw, 4) if gross_margin_raw != '-' else '-'

            cross_proportion = round(fifteen_day_inventory_turnover_rate_raw * gross_margin_raw,
                                     4) if fifteen_day_inventory_turnover_rate_raw != '-' and gross_margin_raw != '-' else '-'

            # 复制到目标文件，并在D列和F列之间插入空白列
            target_sheet.append(
                [operate, data_b, data_d, data_m, start_selling_date, data_f, data_z, data_aa, data_ac, data_ad,
                 data_ak,
                 available_stock,
                 average_cost,
                 end_value,
                 Seven_day_sales,
                 fifteen_day_sales,
                 thirty_day_sales,
                 last_week_seven_day_sales,
                 sales_difference_between_two_weeks_and_seven_days,
                 available_days_for_sale_t30,
                 sell_out_indicator_within_30_days,
                 last_week_inventory_indicators,
                 comparison_between_this_week_and_last_week,
                 inventory_alarm_in_the_past_7_days,
                 inventory_alarm_over_45_days,
                 seven_day_advertisement_cost,
                 fifteen_day_advertisement_cost,
                 thirty_day_advertisement_cost,
                 seven_day_acos,
                 last_week_seven_day_ACOS,
                 two_week_ACOS_difference,
                 fifteen_day_acos,
                 thirty_day_acos,
                 seven_day_inventory_turnover_rate,
                 fifteen_day_inventory_turnover_rate,
                 gross_margin,
                 cross_proportion,
                 seven_day_sales_cost,
                 fifteen_day_sales_cost,
                 site
                 ])

    # 保存目标文件
    target_workbook.save(new_file)
    deleteRow(new_file, 4, "亚马逊-小满1店_US", 2)


def copyArrangeFBA():
    # 获取当前日期时间
    # current_datetime = datetime.now()
    #
    # # 将日期时间格式化为指定格式
    # formatted_datetime = current_datetime.strftime("%Y%m%d%H%M%S")

    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出2"

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "fba库存-new.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)

    # 打开源文件
    source_file = os.path.join(current_directory, folder_name, "Arrange_fba库存.xlsx")
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook.active

    # 打开目标文件
    target_file = "模板-亚马逊库存分析.xlsx"
    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet = target_workbook.active
    target_sheet.title = "亚马逊库存分析"

    # 复制数据
    start_row = 4
    start_column = 1

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet.cell(row=start_row, column=start_column).value = cell.value
            target_sheet.cell(row=start_row, column=start_column).font = Font(name='宋体', size=11)

            target_sheet.cell(row=start_row, column=start_column).alignment = Alignment(vertical='center',
                                                                                        horizontal='center')
            target_sheet.cell(row=start_row, column=start_column).border = border
            if start_column in [29, 30, 31, 32, 33, 36, 37]:
                target_sheet.cell(row=start_row, column=start_column).number_format = '0.00%'

            start_column += 1
        start_row += 1
        start_column = 1  # 重置列号B对应索引2

    rule1 = IconSetRule('3Arrows', "num", [-1, 0, 1], showValue=True, reverse=False)
    rule2 = IconSetRule('3Arrows', "num", [0, 0, 0.0000000001], showValue=True, reverse=False)
    rule3 = IconSetRule('3Arrows', "num", [0, 0, 0], showValue=True, reverse=False)

    target_sheet.conditional_formatting.add(f'S4:S{target_sheet.max_row}', rule1)
    target_sheet.conditional_formatting.add(f'W4:W{target_sheet.max_row}', rule2)
    target_sheet.conditional_formatting.add(f'AE4:AE{target_sheet.max_row}', rule3)

    # 冻结前3行
    target_sheet.freeze_panes = 'A4'

    # 设置第一行的行高为 150 磅
    target_sheet.row_dimensions[1].height = 150

    # 保存目标文件
    target_workbook.save(new_file)


def summary():
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name1 = "输出"
    folder_name2 = "输出2"
    folder_name3 = "输出3"

    wb = load_workbook(os.path.join(current_directory, folder_name2, "fba库存-new.xlsx"))
    # 获取第一个工作簿
    ws = wb.active
    ws.title = "亚马逊库存分析"

    wb1 = load_workbook(os.path.join(current_directory, folder_name1, "汇总.xlsx"))
    # 获取第一个工作簿
    ws1 = wb1.active

    # 加载成本头程.xlsx文件
    wb2 = load_workbook('品牌广告明细sku.xlsx')
    # 获取第一个工作簿
    ws2_1 = wb2.worksheets[0]
    # 获取第二个工作簿
    ws2_2 = wb2.worksheets[1]

    wb3 = load_workbook("亚马逊产品开始出售日期.xlsx")
    # 获取第一个工作簿
    ws3 = wb3.active

    wb4 = load_workbook("成本头程.xlsx")
    # 获取第一个工作簿
    ws4 = wb4.active

    new_ws = wb.create_sheet(title='汇总')
    for row in ws1.iter_rows(values_only=True):
        new_ws.append(row)

    new_ws = wb.create_sheet(title='品牌广告')
    for row in ws2_1.iter_rows(values_only=True):
        new_ws.append(row)

    new_ws = wb.create_sheet(title='币种汇率')
    for row in ws2_2.iter_rows(values_only=True):
        new_ws.append(row)

    new_ws = wb.create_sheet(title='亚马逊产品开始出售日期')
    for row in ws3.iter_rows(values_only=True):
        new_ws.append(row)

    new_ws = wb.create_sheet(title='成本头程')
    for row in ws4.iter_rows(values_only=True):
        new_ws.append(row)

    wb.save(os.path.join(current_directory, folder_name3, '亚马逊库存分析.xlsx'))


if __name__ == '__main__':
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    createDirectory("输出2")
    createDirectory("输出3")
    arrangeCostHeadProcess_Site()
    vlookupCostHeadProcess_Site()
    arrangeFBA()
    copyArrangeFBA()
    summary()
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))