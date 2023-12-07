import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import pandas as pd

from utils import deleteRow, unmergeCells, deleteCells, get_ad_sku_dict, getNumberDaysBetweenDates, currencyConverter

import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def createDirectory():
    # 获取当前工作目录
    current_directory = os.getcwd()

    # 新建文件夹
    folder_name = "输出"
    folder_path = os.path.join(current_directory, folder_name)
    os.makedirs(folder_path)


def addNewColumn():
    # 获取当前工作目录
    current_directory = os.getcwd()

    # 新建文件夹
    folder_name = "输出"
    # 获取当前目录下的所有Excel文件
    files = [file for file in os.listdir() if file.endswith(
        '.xlsx') and file != '品牌广告明细sku.xlsx' and file != '成本头程.xlsx' and file != '模板-亚马逊库存分析.xlsx' and not file.startswith(
        "亚马逊") and not file.startswith("fba")]

    error_files = []
    pattern = r"^(销售|商品推广|展示推广|品牌推广)\d{2}\.\d{2}-\d{2}\.\d{2}$"
    for file in files:
        # 加载Excel文件
        wb = load_workbook(file)
        for sheet in wb.sheetnames:
            if not re.match(pattern, sheet):
                error_files.append(file + sheet)

    # print(error_files)
    if error_files:
        raise ValueError(str(error_files) + "工作簿命名有误")

    for file in files:
        # 加载Excel文件
        wb = load_workbook(file)
        for sheet in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheet]
            # 在第一列的前面插入一列
            ws.insert_cols(1)
            # 设置第一行的值为"店铺"
            ws.cell(row=1, column=1).value = "店铺"
            # 获取表名的前两个字符
            shop_name = file.split("店")[0] + "店" + sheet

            # 设置余下行的值为表名的前两个字符
            i = ws.max_row
            real_max_row = 0
            while i > 0:
                row_dict = {i.value for i in ws[i]}
                if row_dict == {None}:
                    i = i - 1
                else:
                    real_max_row = i
                    break

            for row in range(2, real_max_row + 1):
                ws.cell(row=row, column=1).value = shop_name

            if sheet.startswith("销售"):
                date_diff = getNumberDaysBetweenDates(sheet, "销售")
                if date_diff == 14:
                    # 选择要操作的工作表
                    # ws = wb[sheet]  # 将'Sheet1'替换为您的工作表名称

                    # 查找列名为"已订购商品销售额"的列
                    column_name = "已订购商品销售额"
                    column_index = None
                    for col in ws.iter_cols():
                        if col[0].value == column_name:
                            column_index = col[0].column
                            break

                    # 如果找到了该列，则在其右边插入空白列
                    if column_index is not None:
                        new_column_index = column_index + 1
                        ws.insert_cols(new_column_index)

                        # 设置新列的列名为"换算为人民币"
                        new_column_name = "换算为人民币"
                        ws.cell(row=1, column=new_column_index).value = new_column_name

                        # 获取该列数据的行数
                        num_rows = ws.max_row

                        # 从第二行开始，将新列的值设置为对应行的"已订购商品销售额"列的值加1
                        for row in range(2, num_rows + 1):
                            cell_old = ws.cell(row=row, column=column_index)

                            new_cell_value = currencyConverter(cell_old)

                            ws.cell(row=row, column=new_column_index).value = new_cell_value

            # 构建新的文件名
        output_filename = file.replace('.xlsx', '-new.xlsx')
        new_file = os.path.join(current_directory, folder_name, output_filename)
        # 保存修改后的Excel文件为新文件
        wb.save(new_file)


def mergeFilesByWorkbookName():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    # 获取当前目录下所有以"-new.xlsx"结尾的Excel文件
    files = [file for file in os.listdir(folder_name) if file.endswith('-new.xlsx')]
    output_filename = "mergeFilesByWorkbookName.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)

    # 字典用于存储工作簿名称和对应的工作表数据
    workbook_data = {}

    # 遍历Excel文件
    for file in files:
        # 加载Excel文件
        wb = load_workbook(os.path.join(current_directory, folder_name, file))
        # 遍历工作表
        for sheetname in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheetname]

            # 如果工作簿名称已存在于字典中，则将当前工作表数据追加到已存在的列表中
            if sheetname in workbook_data:
                workbook_data[sheetname].extend(ws.iter_rows(values_only=True))
            else:
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
                workbook_data[sheetname] = list(ws.iter_rows(values_only=True))

    # 创建新的Workbook对象
    merged_wb = Workbook()

    # 遍历工作簿数据字典
    for sheetname, data in workbook_data.items():
        # 创建新的工作表
        ws = merged_wb.create_sheet(title=sheetname)

        # 写入数据到工作表
        for row_index, row_data in enumerate(data):
            for col_index, cell_value in enumerate(row_data):
                column_letter = get_column_letter(col_index + 1)
                ws[column_letter + str(row_index + 1)].value = cell_value

    # 删除默认创建的Sheet
    del merged_wb['Sheet']

    # 保存合并后的Excel文件
    merged_wb.save(new_file)

    deleteRow(new_file, 3, '店铺')
    deleteRow(new_file, 1)


def salesPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(os.path.join(current_directory, folder_name, filename))

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "Merge_SalesPivotTable.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    writer = pd.ExcelWriter(new_file, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("销售"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values='已订购商品数量', index=['店铺', 'SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(new_file)
    deleteRow(new_file, 2)


def fifteenDaySalesPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(os.path.join(current_directory, folder_name, filename))

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "Merge_FifteenDaySalesPivotTable.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    writer = pd.ExcelWriter(new_file, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("销售"):
            date_diff = getNumberDaysBetweenDates(sheet_name, "销售")

            if date_diff == 14:
                # 读取工作簿数据
                df = excel_file.parse(sheet_name)

                # 创建数据透视表
                pivot_table = pd.pivot_table(df, values=['换算为人民币'], index=['店铺', 'SKU', '已订购商品销售额'],
                                             aggfunc='sum')

                # 将数据透视表写入新的工作簿
                pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

                # 获取工作簿的worksheet对象
                worksheet = writer.sheets[sheet_name]

                # 设置数据透视表布局和打印设置
                worksheet.sheet_view.showGridLines = False
                worksheet.sheet_properties.outlinePr.summaryBelow = True
                worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(new_file)
    deleteRow(new_file, 2)


def productPromotionPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(os.path.join(current_directory, folder_name, filename))

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_1_ProductPromotion.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    writer = pd.ExcelWriter(new_file, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("商品推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '7天总销售额'], index=['店铺', '广告SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(new_file)


def displayPromotionPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(os.path.join(current_directory, folder_name, filename))

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_2_DisplayPromotion.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    writer = pd.ExcelWriter(new_file, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("展示推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '14天总销售额'], index=['店铺', '广告SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(new_file)


def brandPromotionPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(os.path.join(current_directory, folder_name, filename))

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_3_BrandPromotion.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    writer = pd.ExcelWriter(new_file, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("品牌推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '14天总销售额'], index=['店铺', '广告活动名称'],
                                         aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(new_file)
    deleteRow(new_file, 1)


def modify_brandPromotionPivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    ad_sku_dict = get_ad_sku_dict("品牌广告明细sku.xlsx", "品牌广告", 0, 1)

    # 打开PivotTable_3_BrandPromotion.xlsx文件
    loadFile = os.path.join(current_directory, folder_name, "PivotTable_3_BrandPromotion.xlsx")
    wb = load_workbook(loadFile)

    # 遍历每个工作表
    for sheet_name in wb.sheetnames:
        # 获取当前工作表
        ws = wb[sheet_name]

        # 获取除第一行外的所有行数据
        rows = list(ws.iter_rows(min_row=2))

        # 遍历每一行数据，向下新增两行
        for row in rows:
            # 获取当前行的行号
            row_number = row[0].row
            date = row[0].value
            name = row[1].value
            total_sales = row[2].value
            cost = row[3].value

            for i in ad_sku_dict:
                if name == i:
                    add_row = len(ad_sku_dict[name])

                    ws.insert_rows(row_number + 1, amount=add_row)

                    for i in range(1, add_row + 1):
                        # print(i)
                        ws.cell(row=row_number + i, column=1, value=date)
                        ws.cell(row=row_number + i, column=2, value=ad_sku_dict[name][i - 1])
                        ws.cell(row=row_number + i, column=3, value=total_sales / add_row)
                        ws.cell(row=row_number + i, column=4, value=cost / add_row)

            # 删除原始行
            ws.delete_rows(row_number, amount=1)
    # 保存修改后的Excel文件
    wb.save(loadFile)


def twoWeeksEndValue():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 新建一个Excel文件用于存储提取的数据

    files = [file for file in os.listdir() if file.startswith('亚马逊库存分析')]
    for file in files:
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_filename = "Merge_EndValue_" + file
        new_file = os.path.join(current_directory, folder_name, output_filename)

        # 加载Excel文件
        wb = load_workbook(file, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet.endswith("库存分析"):
                worksheet = wb[sheet]

                # 复制B列和M列的数据到新的Excel表
                for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=14, values_only=True):
                    output_sheet.append([row[0], row[1], row[12]])

        output_sheet.insert_rows(1)  # 在第一行插入新行
        # 填充标题行
        output_sheet["A1"] = "店铺"
        output_sheet["B1"] = "Seller SKU"
        output_sheet["C1"] = file.split("亚马逊库存分析")[1].split(".")[0] + "期末货值"
        output_workbook.save(new_file)


def twoWeeksEndValueEveryShop():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 新建一个Excel文件用于存储提取的数据

    files = [file for file in os.listdir() if file.startswith('亚马逊库存分析')]



    for file in files:
        shop_and_salesperson = get_ad_sku_dict(file,"亚马逊库存分析", 1 ,0)
        filtered_shop_and_salesperson = {key: value for key, value in shop_and_salesperson.items() if key is not None and key.endswith('店')}

        filtered_dict = {}
        for key, value in filtered_shop_and_salesperson.items():
            filtered_values = [v for v in value if v != '-']
            filtered_dict[key] = len(filtered_values)

        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "店铺SKU数量表"

        # 设置列名
        output_sheet["A1"] = "店铺"
        output_sheet["B1"] = file.split("亚马逊库存分析")[1].split(".")[0] + "总SKU数量"

        # 初始化行数
        row_number = 2

        output_filename = "Merge_EveryShop_" + file
        new_file = os.path.join(current_directory, folder_name, output_filename)

        # 加载Excel文件
        # wb = load_workbook(file, read_only=True, data_only=True)
        for key, value in filtered_dict.items():
            # if sheet.endswith("店"):
            #     worksheet = wb[sheet]

                # 获取工作簿名字和B1数据
                shop_name = key
                sku_count = value

                # 添加数据到新的Excel表
                output_sheet["A{}".format(row_number)] = shop_name
                output_sheet["B{}".format(row_number)] = sku_count

                # 增加行数
                row_number += 1
        output_workbook.save(new_file)


def lastWeekSevenDaySales():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 新建一个Excel文件用于存储提取的数据

    files = [file for file in os.listdir() if file.startswith('亚马逊库存分析')]
    for file in files:
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_filename = "Merge_LastWeekSevenDaySales_" + file
        new_file = os.path.join(current_directory, folder_name, output_filename)

        # 加载Excel文件
        wb = load_workbook(file, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet.endswith("库存分析"):
                worksheet = wb[sheet]

                # 复制B列和M列的数据到新的Excel表
                for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=15, values_only=True):
                    output_sheet.append([row[0], row[1], row[13]])

        output_sheet.insert_rows(1)  # 在第一行插入新行
        # 填充标题行
        output_sheet["A1"] = "店铺"
        output_sheet["B1"] = "Seller SKU"
        output_sheet["C1"] = file.split("亚马逊库存分析")[1].split(".")[0] + "-7天销量"
        output_workbook.save(new_file)


def lastWeekInventoryIndicators():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 新建一个Excel文件用于存储提取的数据

    files = [file for file in os.listdir() if file.startswith('亚马逊库存分析')]
    for file in files:
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_filename = "Merge_LastWeekInventoryIndicators_" + file
        new_file = os.path.join(current_directory, folder_name, output_filename)

        # 加载Excel文件
        wb = load_workbook(file, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet.endswith("库存分析"):
                worksheet = wb[sheet]

                # 复制B列和M列的数据到新的Excel表
                for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=21, values_only=True):
                    output_sheet.append([row[0], row[1], row[19]])

        output_sheet.insert_rows(1)  # 在第一行插入新行
        # 填充标题行
        output_sheet["A1"] = "店铺"
        output_sheet["B1"] = "Seller SKU"
        output_sheet["C1"] = file.split("亚马逊库存分析")[1].split(".")[0] + "-库存指标"
        output_workbook.save(new_file)


def lastWeekSevenDayACOS():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"
    # 新建一个Excel文件用于存储提取的数据

    files = [file for file in os.listdir() if file.startswith('亚马逊库存分析')]
    for file in files:
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_filename = "Merge_LastWeekSevenDayACOS_" + file
        new_file = os.path.join(current_directory, folder_name, output_filename)

        # 加载Excel文件
        wb = load_workbook(file, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet.endswith("库存分析"):
                worksheet = wb[sheet]

                # 复制B列和M列的数据到新的Excel表
                for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=29, values_only=True):
                    output_sheet.append([row[0], row[1], row[27]])

        output_sheet.insert_rows(1)  # 在第一行插入新行
        # 填充标题行
        output_sheet["A1"] = "店铺"
        output_sheet["B1"] = "Seller SKU"
        output_sheet["C1"] = file.split("亚马逊库存分析")[1].split(".")[0] + "-7天ACOS"
        output_workbook.save(new_file)


def mergePivotTable():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    files = [file for file in os.listdir(folder_name) if file.startswith('PivotTable_')]

    output_filename = "Merge_PromotionPivotTable.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)
    # 字典用于存储工作簿名称和对应的工作表数据
    workbook_data = {}

    # 遍历Excel文件
    for file in files:
        # 加载Excel文件
        wb = load_workbook(os.path.join(current_directory, folder_name, file))

        # 遍历工作表
        for sheetname in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheetname]

            # 如果工作簿名称已存在于字典中，则将当前工作表数据追加到已存在的列表中
            if sheetname[4:] in workbook_data:
                workbook_data[sheetname[4:]].extend(ws.iter_rows(values_only=True))
            else:
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
                workbook_data[sheetname[4:]] = list(ws.iter_rows(values_only=True))

    # 创建新的Workbook对象
    merged_wb = Workbook()

    # 遍历工作簿数据字典
    for sheetname, data in workbook_data.items():
        # 创建新的工作表
        ws = merged_wb.create_sheet(title=sheetname)

        # 写入数据到工作表
        for row_index, row_data in enumerate(data):
            for col_index, cell_value in enumerate(row_data):
                column_letter = get_column_letter(col_index + 1)
                ws[column_letter + str(row_index + 1)].value = cell_value

    # 删除默认创建的Sheet
    del merged_wb['Sheet']

    # 保存合并后的Excel文件
    merged_wb.save(new_file)

    deleteRow(new_file, 1)
    deleteRow(new_file, 3, '店铺')


def summary():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name = "输出"

    # 获取文件夹中以"Merge_"开头的所有Excel文件
    files = [file for file in os.listdir(folder_name) if file.startswith('Merge_')]
    output_filename = "汇总.xlsx"
    new_file = os.path.join(current_directory, folder_name, output_filename)

    # 创建一个空的DataFrame用于存储合并后的数据
    merged_data = pd.DataFrame()

    # 遍历每个Excel文件并将所有工作簿横向合并
    for file in files:
        # file_path = os.path.join(folder_path, file)

        # 读取Excel文件中的所有工作簿
        xl = pd.ExcelFile(os.path.join(current_directory, folder_name, file))
        sheet_names = xl.sheet_names

        # 遍历每个工作簿并将其横向合并到merged_data中
        for sheet_name in sheet_names:
            # 读取当前工作簿的数据
            df = pd.read_excel(os.path.join(current_directory, folder_name, file), sheet_name=sheet_name)

            # 在最右侧新增两个空白列
            df_with_columns = df.assign(NewColumn1="", NewColumn2="")

            # 将数据横向合并到merged_data中
            merged_data = pd.concat([merged_data, df_with_columns], axis=1)

    # 将合并后的数据写入新的Excel文件
    merged_data.to_excel(new_file, sheet_name='汇总', index=False)

    # 删除指定单元格
    deleteCells(new_file, "NewColumn")


if __name__ == '__main__':
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    createDirectory()
    addNewColumn()
    mergeFilesByWorkbookName()
    salesPivotTable()
    fifteenDaySalesPivotTable()
    productPromotionPivotTable()
    displayPromotionPivotTable()
    brandPromotionPivotTable()
    modify_brandPromotionPivotTable()
    twoWeeksEndValue()
    twoWeeksEndValueEveryShop()
    lastWeekSevenDaySales()
    lastWeekInventoryIndicators()
    lastWeekSevenDayACOS()
    mergePivotTable()
    summary()

    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))