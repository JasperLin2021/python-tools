import json
import os

from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import pyautogui


# 生成xlsx文件及数据错误及数据缺少情况文件
def determineDailyBills(current_directory, month_content):
    # 从文件中读取JSON字符串并解析为字典
    with open('各店铺当月每天回款情况.txt', 'r', encoding='utf-8') as file:
        monthlyPaymentDetails_dict = json.load(file)

    # processed_csv = {'数据正确的文件':{}}
    lack_files = {'缺少的日期': {}}
    error_files = {'数据错误的文件': {}}

    if not os.path.isfile("数据错误及数据缺少情况.txt") or os.path.getsize("数据错误及数据缺少情况.txt") == 0:
        folder_list = [folder for folder in os.listdir('.') if os.path.isdir(folder) and folder != '__pycache__' and folder != '输出']
        for folder in folder_list:
            directory = os.path.join(current_directory, folder)
            # 处理未转化的csv文件

            all_files_list = [item.split('.csv')[0].split('.')[1] for item in os.listdir(directory) if
                              item.endswith(".csv")]
            monthlyPaymentDetails_list = list(monthlyPaymentDetails_dict[folder].keys())
            difference = list(set(monthlyPaymentDetails_list) - set(all_files_list))

            if difference != []:
                for i in difference:
                    if lack_files.get('缺少的日期') != None and lack_files.get('缺少的日期').get(
                            folder) != None:
                        lack_files.get('缺少的日期').get(folder).append(str(month_content) + '.' + str(i) + '.csv')
                    else:
                        lack_files['缺少的日期'][folder] = [str(month_content) + '.' + str(i) + '.csv']

            else:
                for filename in os.listdir(directory):
                    file_path = os.path.join(directory, filename)
                    # 将csv文件转化为xlsx文件，并格式化数据
                    if file_path.endswith(".csv"):
                        simple_filename = filename.split(".csv")[0]
                        new_file_path = os.path.join(directory, simple_filename + ".xlsx")

                        # # 读取CSV文件
                        f = open(file_path, 'r', encoding='utf-8')
                        # 创建一个workbook 设置编码
                        workbook = Workbook()
                        # 创建一个worksheet
                        worksheet = workbook.active
                        workbook.title = 'sheet'

                        for line in f:
                            if '","' in line:
                                new_line = line.replace('","', '^&').replace('"', '').replace('\n', '')
                                new_line = new_line.split('^&')
                                worksheet.append(new_line)
                            else:
                                new_line = line.replace('"', '').replace('\n', '')

                                worksheet.append([new_line])

                        # 获取K列中最后一个有数据的单元格的行数
                        last_row = worksheet.max_row

                        # 遍历K列中的每个单元格
                        total_list = []
                        column_list = [11, 27, 28]
                        cell_list = ['K10', 'AA10', 'AB10']

                        for c in column_list:
                            total = 0
                            for row in range(12, last_row + 1):
                                cell = worksheet.cell(row=row, column=c)
                                if cell.value == '--':
                                    continue  # K列的列号是11
                                # print(folder+simple_filename)
                                cell.value = float(cell.value)  # 将单元格的数据类型设置为数字
                                total += float(cell.value)
                            total_list.append(total)
                        #
                        if float(worksheet['B10'].value.split()[0].replace(",", "")) == float(
                                monthlyPaymentDetails_dict.get(folder).get(simple_filename.split('.')[1])):
                            result_dict = {key: value for key, value in zip(cell_list, total_list)}
                            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            for k, v in result_dict.items():
                                worksheet[k].value = v
                                worksheet[k].fill = yellow_fill

                            workbook.save(new_file_path)
                        else:
                            if error_files.get('数据错误的文件') != None and error_files.get('数据错误的文件').get(
                                    folder) != None:
                                error_files.get('数据错误的文件').get(folder).append(filename)
                            else:
                                error_files['数据错误的文件'][folder] = [filename]

        print(error_files)
        print(lack_files)
        with open('数据错误及数据缺少情况.txt', 'w', encoding='utf-8') as file:
            json.dump(error_files, file, ensure_ascii=False)
            file.write('\n')
            json.dump(lack_files, file, ensure_ascii=False)

    else:
        error_dict = {}
        lack_dict = {}
        with open('数据错误及数据缺少情况.txt', 'r', encoding='utf-8') as file:
            for line in file:
                dict = json.loads(line)
                if dict.get('数据错误的文件'):
                    error_dict = dict.get('数据错误的文件')
                else:
                    lack_dict = dict.get('缺少的日期')

            # 合并字典A和字典B生成新的字典
        merged_dict = {}
        merged_dict.update(error_dict)

        for key, value in lack_dict.items():
            if key in merged_dict:
                merged_dict[key] += value
            else:
                merged_dict[key] = value

        for key, value in merged_dict.items():
            directory = os.path.join(current_directory, key)

            # 判断文件是否缺失
            all_files_list = [item.split('.csv')[0].split('.')[1] for item in os.listdir(directory) if
                              item.endswith(".csv")]
            monthlyPaymentDetails_list = list(monthlyPaymentDetails_dict[key].keys())
            difference = list(set(monthlyPaymentDetails_list) - set(all_files_list))

            if difference != []:
                for i in difference:
                    if lack_files.get('缺少的日期') != None and lack_files.get('缺少的日期').get(
                            key) != None:
                        lack_files.get('缺少的日期').get(key).append(str(month_content) + '.' + str(i) + '.csv')
                    else:
                        lack_files['缺少的日期'][key] = [str(month_content) + '.' + str(i) + '.csv']

            else:
                # 生成汇总
                for filename in os.listdir(directory):
                    file_path = os.path.join(directory, filename)

                    if file_path.endswith(".csv"):
                        simple_filename = filename.split(".csv")[0]
                        new_file_path = os.path.join(directory, simple_filename + ".xlsx")

                        # # 读取CSV文件
                        f = open(file_path, 'r', encoding='utf-8')
                        # 创建一个workbook 设置编码
                        workbook = Workbook()
                        # 创建一个worksheet
                        worksheet = workbook.active
                        workbook.title = 'sheet'

                        for line in f:
                            if '","' in line:
                                new_line = line.replace('","', '^&').replace('"', '').replace('\n', '')
                                new_line = new_line.split('^&')
                                worksheet.append(new_line)
                            else:
                                new_line = line.replace('"', '').replace('\n', '')

                                worksheet.append([new_line])

                        # 获取K列中最后一个有数据的单元格的行数
                        last_row = worksheet.max_row

                        # 遍历K列中的每个单元格
                        total_list = []
                        column_list = [11, 27, 28]
                        cell_list = ['K10', 'AA10', 'AB10']

                        for c in column_list:
                            total = 0
                            for row in range(12, last_row + 1):
                                cell = worksheet.cell(row=row, column=c)
                                if cell.value == '--':
                                    continue  # K列的列号是11
                                cell.value = float(cell.value)  # 将单元格的数据类型设置为数字
                                total += float(cell.value)
                            total_list.append(total)

                        if float(worksheet['B10'].value.split()[0]) == float(
                                monthlyPaymentDetails_dict.get(key).get(simple_filename.split('.')[1])):
                            result_dict = {key: value for key, value in zip(cell_list, total_list)}
                            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            for k, v in result_dict.items():
                                worksheet[k].value = v
                                worksheet[k].fill = yellow_fill

                            workbook.save(new_file_path)
                        else:
                            if error_files.get('数据错误的文件') != None and error_files.get('数据错误的文件').get(
                                    key) != None:
                                error_files.get('数据错误的文件').get(key).append(filename)
                            else:
                                error_files['数据错误的文件'][key] = [filename]

        print(error_files)
        print(lack_files)
        with open('数据错误及数据缺少情况.txt', 'w', encoding='utf-8') as file:
            json.dump(error_files, file, ensure_ascii=False)
            file.write('\n')
            json.dump(lack_files, file, ensure_ascii=False)


def createDirectory(current_directory):
    # 新建文件夹
    folder_name = "输出"
    folder_path = os.path.join(current_directory, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        return folder_path
    return folder_path


def mergeFiles(current_directory, new_directory, big_dict, select_refund):
    error_dict = {}
    lack_dict = {}
    with open('数据错误及数据缺少情况.txt', 'r', encoding='utf-8') as file:
        for line in file:
            dict = json.loads(line)
            if dict.get('数据错误的文件'):
                error_dict = dict.get('数据错误的文件')
            else:
                lack_dict = dict.get('缺少的日期')

        # 合并字典A和字典B生成新的字典
    merged_dict = {}
    merged_dict.update(error_dict)

    for key, value in lack_dict.items():
        if key in merged_dict:
            merged_dict[key] += value
        else:
            merged_dict[key] = value

    image_extensions = [".jpg", ".jpeg", ".png", ".gif", ".bmp"]
    c_cell = eval(select_refund.get('c_cell'))
    ak_cell = eval(select_refund.get('ak_cell'))
    for k, v in big_dict.items():
        if merged_dict.get(k):
            continue
        merged_wb = Workbook()
        workbook_all_data = {}
        incomeStatistics = []
        picture_file_list = [os.path.abspath(os.path.join(current_directory, k, picture_file)) for picture_file in
                             os.listdir('./' + k) if
                             any(picture_file.lower().endswith(ext) for ext in image_extensions)]
        for file in v:
            workbook_data = {}
            wb = load_workbook(file)
            ws = wb.active
            # 在第一列的前面插入一列
            ws.insert_cols(1)
            # 设置第一行的值为"店铺"
            ws.cell(row=1, column=1).value = "date"
            # 设置余下行的值为表名的前两个字符

            max_row = ws.max_row

            sheet_name = file.split(k)[1].split('\\')[1].split('.xlsx')[0]
            for row in range(2, max_row + 1):
                ws.cell(row=row, column=1).value = sheet_name
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
            workbook_data[k] = list(ws.iter_rows(values_only=True))

            net_income = round(ws['L10'].value, 2)
            fixed_payment_fee = round(ws['AB10'].value, 2)
            floating_transaction_fee = round(ws['AC10'].value, 2)
            if k in workbook_all_data and sheet_name.split('.')[1] != 1:
                ws.delete_rows(1, 11)
                workbook_all_data[k].extend(ws.iter_rows(values_only=True))
                # workbook_all_data[k].extend(ws.iter_rows(values_only=True))
            else:
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
                workbook_all_data[k] = list(ws.iter_rows(values_only=True))

            # 在汇总表中创建每一个分表
            merged_ws = merged_wb.create_sheet(title=sheet_name)
            for sheetname, data in workbook_data.items():
                # 创建新的工作表
                # 写入数据到工作表
                for row_index, row_data in enumerate(data):
                    for col_index, cell_value in enumerate(row_data):
                        column_letter = get_column_letter(col_index + 1)
                        merged_ws[column_letter + str(row_index + 1)].value = cell_value

            # 获取分表的收入费用统计表明细
            index = sheet_name.split('.')[1]
            date = sheet_name.split('.')[0] + '月' + index + '日'

            record = [int(index), date, net_income, fixed_payment_fee, floating_transaction_fee]
            incomeStatistics.append(record)

        # 1、保存合并后的Excel文件
        merged_all_ws = merged_wb.create_sheet(title='总表')
        for sheetname, data in workbook_all_data.items():
            # 创建新的工作表
            # 写入数据到工作表
            for row_index, row_data in enumerate(data):
                for col_index, cell_value in enumerate(row_data):
                    column_letter = get_column_letter(col_index + 1)
                    merged_all_ws[column_letter + str(row_index + 1)].value = cell_value

            # 删除默认创建的Sheet
        del merged_wb['Sheet']

        ws_totalTable = merged_wb['总表']

        # 2、总表格式化数据
        total_list = []
        column_list = [12, 28, 29]
        cell_list = ['L10', 'AB10', 'AC10']

        for c in column_list:
            total = 0
            for row in range(12, ws_totalTable.max_row + 1):
                cell = ws_totalTable.cell(row=row, column=c)
                if cell.value == '--':
                    continue  # K列的列号是11
                cell.value = float(cell.value)  # 将单元格的数据类型设置为数字
                total += float(cell.value)
            total_list.append(round(total, 2))

        result_dict = {key: value for key, value in zip(cell_list, total_list)}
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for result_dict_key, result_dict_value in result_dict.items():
            ws_totalTable[result_dict_key].value = result_dict_value
            ws_totalTable[result_dict_key].fill = yellow_fill

        # 3、制作退款累计总明细
        ws_totalRefundCumulativeDetails = merged_wb.create_sheet(title='退款累计总明细')
        # c_cell = ['Type', 'Refund', '退款', 'Claim', '索赔', 'Payment dispute', '付款纠纷']
        # ak_cell = ['Cancel', '取消']
        num = 0
        for select_cell_key in c_cell:
            for row in ws_totalTable.iter_rows(min_row=11, values_only=True):
                cell_value_c = row[2]  # 获取B列的值（索引从0开始）
                cell_value_ak = row[36]  # 获取B列的值（索引从0开始）
                if cell_value_c == select_cell_key and not any(item in cell_value_ak for item in ak_cell):
                    num += 1
                    # 将满足条件的行复制到目标工作簿
                    ws_totalRefundCumulativeDetails.append(row)
        # 在第一行之前插入空白行
        ws_totalRefundCumulativeDetails.insert_rows(1)
        ws_totalRefundCumulativeDetails['D1'].value = num - 1
        ws_totalRefundCumulativeDetails['D1'].fill = yellow_fill

        # 4、制作收入费用统计表
        pink_fill = PatternFill(start_color="FED6D3", end_color="FED6D3", fill_type="solid")
        red_font = Font(color="FF0000")
        ws_incomeStatisticsTable = merged_wb.create_sheet(title='收入费用统计表')
        sum_net_income = sum_fixed_payment_fee = sum_floating_transaction_fee = 0
        for r in incomeStatistics:
            ws_incomeStatisticsTable.append(r)
            sum_net_income += r[2]
            sum_fixed_payment_fee += r[3]
            sum_floating_transaction_fee += r[4]
        # 在第一行之前插入空白行
        ws_incomeStatisticsTable.insert_rows(1)
        ws_incomeStatisticsTable['B1'].value = '日期'
        ws_incomeStatisticsTable['C1'].value = '净收入'
        ws_incomeStatisticsTable['D1'].value = '固定成交费'
        ws_incomeStatisticsTable['E1'].value = '浮动成交费'
        max_row = ws_incomeStatisticsTable.max_row
        ws_incomeStatisticsTable[f'C{max_row + 1}'].value = round(sum_net_income, 2)
        ws_incomeStatisticsTable[f'C{max_row + 1}'].fill = pink_fill
        ws_incomeStatisticsTable[f'D{max_row + 1}'].value = round(sum_fixed_payment_fee, 2)
        ws_incomeStatisticsTable[f'E{max_row + 1}'].value = round(sum_floating_transaction_fee, 2)
        ws_incomeStatisticsTable[f'C{max_row + 4}'].value = ws_totalTable['L10'].value
        ws_incomeStatisticsTable[f'C{max_row + 4}'].fill = pink_fill
        ws_incomeStatisticsTable[f'D{max_row + 4}'].value = ws_totalTable['AB10'].value
        ws_incomeStatisticsTable[f'E{max_row + 4}'].value = ws_totalTable['AC10'].value
        ws_incomeStatisticsTable[f'C{max_row + 6}'].value = round(sum_net_income, 2) - ws_totalTable['L10'].value
        ws_incomeStatisticsTable[f'D{max_row + 6}'].value = round(sum_fixed_payment_fee, 2) - ws_totalTable[
            'AB10'].value
        ws_incomeStatisticsTable[f'D{max_row + 6}'].font = red_font
        ws_incomeStatisticsTable[f'E{max_row + 6}'].value = round(sum_floating_transaction_fee, 2) - ws_totalTable[
            'AC10'].value
        # print(ws_incomeStatisticsTable.max_row)
        for index, value in enumerate(picture_file_list):
            img = Image(value)  # 图片文件的路径
            ws_incomeStatisticsTable.add_image(img, f'G{index + 1}')  # 将图片插入到A1单元格

        sheet_names = merged_wb.sheetnames
        totalTable_sheet_index = sheet_names.index('总表')
        merged_wb.move_sheet(merged_wb[sheet_names[totalTable_sheet_index]], offset=-(len(sheet_names) - 3))
        incomeStatisticsTable_sheet_index = sheet_names.index('收入费用统计表')
        merged_wb.move_sheet(merged_wb[sheet_names[incomeStatisticsTable_sheet_index]],
                             offset=-(len(sheet_names) - 2))
        totalRefundCumulativeDetails_sheet_index = sheet_names.index('退款累计总明细')
        merged_wb.move_sheet(merged_wb[sheet_names[totalRefundCumulativeDetails_sheet_index]],
                             offset=-(len(sheet_names) - 3))

        new_file = os.path.join(new_directory, k + ".xlsx")
        merged_wb.save(new_file)


if __name__ == '__main__':
    current_directory = os.getcwd()
    file = open('月份.txt', 'r')
    month_content = int(file.read())
    file.close()

    determineDailyBills(current_directory, month_content)

    new_directory = createDirectory(current_directory)
    folder_list = [folder for folder in os.listdir('.') if
                   os.path.isdir(folder) and not folder.startswith('输出') and not folder.startswith('__')]
    # 所有xlsx文件大字典
    big_dict = {}
    for folder in folder_list:
        file_list = []
        directory = os.path.join(current_directory, folder)
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            # 检查文件是否是图片文件
            if os.path.isfile(file_path) and filename.lower().endswith(".xlsx") and not filename.lower().startswith(
                    "~$"):
                file_list.append(file_path)


        def sort_key(file_name):
            # 提取文件名中的数字部分（例如，'9.1.csv' -> ('9', '1', '.csv') -> 9.1）
            numeric_part = file_name.split(folder)[1].split('\\')[1].split('.')[1]
            return float(numeric_part)


        sorted_list = sorted(file_list, key=sort_key)

        big_dict[folder] = sorted_list

    select_refund = {}
    with open('退款筛选条件.txt', 'r', encoding='utf-8') as file:
        for line in file:
            line_dict = json.loads(line)
            select_refund.update(line_dict)

    mergeFiles(current_directory, new_directory, big_dict, select_refund)

    message = "完成！"
    title = "截图识别完成"
    pyautogui.alert(message, title)
