import os

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import pandas as pd

from utils import deleteRow, unmergeCells, deleteCells, get_ad_sku_dict


def addNewColumn():
    # 获取当前目录下的所有Excel文件
    files = [file for file in os.listdir() if file.endswith('.xlsx')]
    for file in files:
        # 加载Excel文件
        wb = load_workbook(file)
        for sheet in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheet]
            # 在第一列的前面插入一列
            ws.insert_cols(1)
            # 设置第一行的值为"店铺"
            ws.cell(row=1, column=1).value = "店铺"
            # 获取表名的前两个字符
            shop_name = file[:2] + sheet

            # 设置余下行的值为表名的前两个字符
            i = ws.max_row
            real_max_row = 0
            while i > 0:
                row_dict = {i.value for i in ws[i]}
                if row_dict == {None}:
                    i = i - 1
                else:
                    real_max_row = i
                    break

            for row in range(2, real_max_row + 1):
                ws.cell(row=row, column=1).value = shop_name
            # 构建新的文件名
        new_file = file.replace('.xlsx', '-new.xlsx')
        # 保存修改后的Excel文件为新文件
        wb.save(new_file)


def mergeFilesByWorkbookName():
    # 获取当前目录下所有以"-new.xlsx"结尾的Excel文件
    files = [file for file in os.listdir() if file.endswith('-new.xlsx')]
    output_filename = "mergeFilesByWorkbookName.xlsx"

    # 字典用于存储工作簿名称和对应的工作表数据
    workbook_data = {}

    # 遍历Excel文件
    for file in files:
        # 加载Excel文件
        wb = load_workbook(file)
        # 遍历工作表
        for sheetname in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheetname]

            # 如果工作簿名称已存在于字典中，则将当前工作表数据追加到已存在的列表中
            if sheetname in workbook_data:
                workbook_data[sheetname].extend(ws.iter_rows(values_only=True))
            else:
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
                workbook_data[sheetname] = list(ws.iter_rows(values_only=True))

    # 创建新的Workbook对象
    merged_wb = Workbook()

    # 遍历工作簿数据字典
    for sheetname, data in workbook_data.items():
        # 创建新的工作表
        ws = merged_wb.create_sheet(title=sheetname)

        # 写入数据到工作表
        for row_index, row_data in enumerate(data):
            for col_index, cell_value in enumerate(row_data):
                column_letter = get_column_letter(col_index + 1)
                ws[column_letter + str(row_index + 1)].value = cell_value

    # 删除默认创建的Sheet
    del merged_wb['Sheet']

    # 保存合并后的Excel文件
    merged_wb.save(output_filename)

    deleteRow(output_filename, 3, '店铺')


def salesPivotTable():
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(filename)

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "Merge_SalesPivotTable.xlsx"
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("销售"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values='已订购商品数量', index=['店铺', 'SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(output_filename)
    deleteRow(output_filename, 2)


def productPromotionPivotTable():
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(filename)

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_1_ProductPromotion.xlsx"
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("商品推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '7天总销售额'], index=['店铺', '广告SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(output_filename)


def displayPromotionPivotTable():
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(filename)

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_2_DisplayPromotion.xlsx"
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("展示推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '14天总销售额'], index=['店铺', '广告SKU'], aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(output_filename)


def brandPromotionPivotTable():
    # 读取Excel文件
    filename = "mergeFilesByWorkbookName.xlsx"
    excel_file = pd.ExcelFile(filename)

    # 获取所有工作簿名称
    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "PivotTable_3_BrandPromotion.xlsx"
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')

    # 遍历每个工作簿
    for sheet_name in sheet_names:
        if sheet_name.startswith("品牌推广"):
            # 读取工作簿数据
            df = excel_file.parse(sheet_name)

            # 创建数据透视表
            pivot_table = pd.pivot_table(df, values=['花费', '14天总销售额'], index=['店铺', '广告活动名称'],
                                         aggfunc='sum')

            # 将数据透视表写入新的工作簿
            pivot_table.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)

            # 获取工作簿的worksheet对象
            worksheet = writer.sheets[sheet_name]

            # 设置数据透视表布局和打印设置
            worksheet.sheet_view.showGridLines = False
            worksheet.sheet_properties.outlinePr.summaryBelow = True
            worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(output_filename)
    deleteRow(output_filename, 1)


def modify_BrandPromotion():
    ad_sku_dict = get_ad_sku_dict()
    # 打开PivotTable_3_BrandPromotion.xlsx文件
    # 打开PivotTable_3_BrandPromotion.xlsx文件
    wb = load_workbook("PivotTable_3_BrandPromotion.xlsx")

    # 遍历每个工作表
    for sheet_name in wb.sheetnames:
        # 获取当前工作表
        ws = wb[sheet_name]

        # 获取除第一行外的所有行数据
        rows = list(ws.iter_rows(min_row=2))

        # 遍历每一行数据，向下新增两行
        for row in rows:
            # 获取当前行的行号
            row_number = row[0].row
            date = row[0].value
            name = row[1].value
            total_sales = row[2].value
            cost = row[3].value

            for i in ad_sku_dict:
                if name == i:
                    add_row = len(ad_sku_dict[name])

                    ws.insert_rows(row_number + 1, amount=add_row)

                    for i in range(1, add_row + 1):
                        # print(i)
                        ws.cell(row=row_number + i, column=1, value=date)
                        ws.cell(row=row_number + i, column=2, value=ad_sku_dict[name][i - 1])
                        ws.cell(row=row_number + i, column=3, value=total_sales / add_row)
                        ws.cell(row=row_number + i, column=4, value=cost / add_row)

            # 删除原始行
            ws.delete_rows(row_number, amount=1)
    # 保存修改后的Excel文件
    wb.save("PivotTable_3_BrandPromotion.xlsx")

    # 打印广告活动名称和对应的广告SKU值
    # for ad_name, ad_skus in ad_sku_dict.items():
    #     print(f"广告活动名称: {ad_name}, 广告SKU: {ad_skus}")


def mergePivotTable():
    # 获取当前目录下所有以"-new.xlsx"结尾的Excel文件
    files = [file for file in os.listdir() if file.startswith('PivotTable_')]
    output_filename = "Merge_PromotionPivotTable.xlsx"
    # 字典用于存储工作簿名称和对应的工作表数据
    workbook_data = {}

    # 遍历Excel文件
    for file in files:
        # 加载Excel文件
        wb = load_workbook(file)

        # 遍历工作表
        for sheetname in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheetname]

            # 如果工作簿名称已存在于字典中，则将当前工作表数据追加到已存在的列表中
            if sheetname[4:] in workbook_data:
                workbook_data[sheetname[4:]].extend(ws.iter_rows(values_only=True))
            else:
                # 如果工作簿名称不存在于字典中，则将当前工作表数据添加到字典中
                workbook_data[sheetname[4:]] = list(ws.iter_rows(values_only=True))

    # 创建新的Workbook对象
    merged_wb = Workbook()

    # 遍历工作簿数据字典
    for sheetname, data in workbook_data.items():
        # 创建新的工作表
        ws = merged_wb.create_sheet(title=sheetname)

        # 写入数据到工作表
        for row_index, row_data in enumerate(data):
            for col_index, cell_value in enumerate(row_data):
                column_letter = get_column_letter(col_index + 1)
                ws[column_letter + str(row_index + 1)].value = cell_value

    # 删除默认创建的Sheet
    del merged_wb['Sheet']

    # 保存合并后的Excel文件
    merged_wb.save(output_filename)

    deleteRow(output_filename, 1)
    deleteRow(output_filename, 3, '店铺')


def summary():
    # 获取文件夹中以"Merge_"开头的所有Excel文件
    files = [file for file in os.listdir() if file.startswith('Merge_')]
    output_filename = "汇总.xlsx"

    # 创建一个空的DataFrame用于存储合并后的数据
    merged_data = pd.DataFrame()

    # 遍历每个Excel文件并将所有工作簿横向合并
    for file in files:
        # file_path = os.path.join(folder_path, file)

        # 读取Excel文件中的所有工作簿
        xl = pd.ExcelFile(file)
        sheet_names = xl.sheet_names

        # 遍历每个工作簿并将其横向合并到merged_data中
        for sheet_name in sheet_names:
            # 读取当前工作簿的数据
            df = pd.read_excel(file, sheet_name=sheet_name)

            # 在最右侧新增两个空白列
            df_with_columns = df.assign(NewColumn1="", NewColumn2="")

            # 将数据横向合并到merged_data中
            merged_data = pd.concat([merged_data, df_with_columns], axis=1)

    # 将合并后的数据写入新的Excel文件
    merged_data.to_excel(output_filename, index=False)

    # 删除指定单元格
    deleteCells(output_filename, "NewColumn")


if __name__ == '__main__':
    print("这个脚本正在直接运行。")

    addNewColumn()
    mergeFilesByWorkbookName()
    salesPivotTable()
    productPromotionPivotTable()
    displayPromotionPivotTable()
    brandPromotionPivotTable()
    modify_BrandPromotion()
    mergePivotTable()
    summary()
