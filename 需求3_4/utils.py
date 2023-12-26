from openpyxl.reader.excel import load_workbook


def unmergeCells(file):
    # 加载Excel文件
    workbook = load_workbook(file)

    # 遍历所有工作簿
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 获取所有合并单元格的范围和值
        merged_cells = list(sheet.merged_cells)  # 创建副本
        merged_values = {}

        # 遍历合并单元格
        for merged_range in merged_cells:
            # 获取合并单元格的范围边界
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

            # 获取合并单元格的值
            merged_value = sheet.cell(row=min_row, column=min_col).value

            # 保存合并单元格的值和范围
            merged_values[merged_range.coord] = {
                'value': merged_value,
                'range': (min_row, min_col, max_row, max_col)
            }

            # 取消合并单元格
            sheet.unmerge_cells(str(merged_range))

        # 恢复取消合并的单元格的值
        for coord, data in merged_values.items():
            merged_value = data['value']
            min_row, min_col, max_row, max_col = data['range']
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = merged_value

    # 保存修改后的Excel文件
    workbook.save(file)

def deleteRow(file, type, keyword=None, column=None):
    wb = load_workbook(file)

    # 删除空白行
    if type == 1:
        # 遍历所有工作表
        for ws in wb.worksheets:
            rows = list(ws.rows)
            for row in reversed(rows):
                if row[0].value is None:
                    ws.delete_rows(row[0].row)
    # 删除第一行
    elif type == 2:
        # 遍历所有工作表
        for ws in wb.worksheets:
            ws.delete_rows(1)

    # 删除第二行及之后的包含"店铺"的行
    elif type == 3:
        for ws in wb.worksheets:
            rows = list(ws.rows)
            for row in rows[1:]:
                if row[0].value is not None and keyword in row[0].value:
                    ws.delete_rows(row[0].row)

    #从最后一行开始遍历每一行，如果某行的某列单元格的值不为空且包含指定关键字（keyword），则删除该行。
    elif type == 4:
        for ws in wb.worksheets:
            for row in reversed(range(1, ws.max_row + 1)):
                if ws.cell(row=row, column=column).value is not None and keyword in ws.cell(row=row, column=column).value:
                    ws.delete_rows(row)
    # 保存修改
    wb.save(file)
