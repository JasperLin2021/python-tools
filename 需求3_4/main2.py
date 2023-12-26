import csv
import os
from datetime import datetime

import openpyxl
import pyautogui
from openpyxl.workbook import Workbook
import pandas as pd

from utils import unmergeCells, deleteRow


def createDirectory(current_directory,output_folder_name):
    # 新建文件夹
    folder_path = os.path.join(current_directory, output_folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        return folder_path
    return folder_path
def merge_bill(bill_folder_name):
    files = [file for file in os.listdir(bill_folder_name) if file.endswith(
        '.csv')]

    df_account_list = []
    for file in files:
        file_path = os.path.join(current_directory, bill_folder_name, file)
        account = file.replace(" ", "").split("账单")[0].replace(".csv", "")

        with open(file_path, encoding='utf-8-sig') as f:
            for i, row in enumerate(csv.reader(f, skipinitialspace=True), 1):
                if i >= 13:
                    row.insert(0, account)
                    row[11] = float(row[11].replace(',','')) if row[11] != '--' else 0
                    # for i in range(len(row)):
                    #     if row[i].isdigit():
                    #         row[i] = float(row[i])
                    df_account_list.append(row)
                    # print(row)

    df_account_list.insert(0, ['账号', '交易创建日期', '类型', '订单编号', '旧订单编号', '买家用户名', '买家姓名', '收货人所在县/市', '运送至省/地区/州', '收货人邮政编码', '收货人所在国家/地区', '净额', '发款货币', '发款日期', '发款编号', '收款方式', '发款状态', '冻结原因', '物品编号', '交易编号', '物品标题', '自定义标签', '数量', '物品小计', '运费与处理费', '卖家收取的税费', 'eBay 收取的税费', '成交费 — 固定', '成交费 — 因品类而异', '“物品与描述不符”指数非常高的费用', '表现不合格的费用', '跨国交易费用', '订金处理费', '交易总金额', '交易货币', '汇率', '参考编号', '描述'])

    # 创建工作簿和工作表
    workbook = Workbook()
    sheet = workbook.active
    #
    # # 将列表数据写入工作表
    for row in df_account_list:
        sheet.append(row)

    sheet.title = "总表"
    new_file = os.path.join(new_directory, str(current_year % 100) + "年" + str(current_month) + "月账单核算.xlsx")
    workbook.save(new_file)

    return new_file
def pivot(bill_file,current_directory,output_folder_name):
    excel_file = pd.ExcelFile(bill_file)

    sheet_names = excel_file.sheet_names

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "Merge_BillPivotTable.xlsx"
    BillPivotTable_file = os.path.join(current_directory, output_folder_name, output_filename)
    writer = pd.ExcelWriter(BillPivotTable_file, engine='openpyxl')

    for sheet_name in sheet_names:

        df = excel_file.parse(sheet_name)

        # 创建数据透视表
        pivot_table = pd.pivot_table(df, values='净额', index=['账号', '类型', '描述'], aggfunc='sum')
        # pivot_table = pd.pivot_table(df, values='净额', index=['账号'], aggfunc='sum')

        # 将数据透视表写入新的工作簿
        pivot_table.to_excel(writer, sheet_name="Sheet1", index=True, startrow=1)

        # 获取工作簿的worksheet对象
        worksheet = writer.sheets["Sheet1"]

        # 设置数据透视表布局和打印设置
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.outlinePr.summaryBelow = True
        worksheet.sheet_properties.outlinePr.summaryRight = True

    # 保存并关闭Excel文件
    writer.save()
    writer.close()

    unmergeCells(BillPivotTable_file)
    deleteRow(BillPivotTable_file, 2)

    return BillPivotTable_file

def merge_all(bill_file, BillPivotTable_file):
    # 加载源文件和目标文件
    source_workbook = openpyxl.load_workbook(BillPivotTable_file)
    target_workbook = openpyxl.load_workbook(bill_file)

    # 获取源文件中的工作簿
    source_sheet = source_workbook['Sheet1']

    # 创建目标文件的新工作簿
    target_sheet = target_workbook.create_sheet('Sheet1')

    # 将源工作簿的内容复制到目标工作簿
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value

    # 保存目标文件
    target_workbook.save(bill_file)

if __name__ == '__main__':
    current_date = datetime.now()
    # 获取当前年份
    current_year = current_date.year
    # 获取当前月份
    current_month = current_date.month
    current_directory = os.getcwd()
    #
    output_folder_name = "输出"
    bill_folder_name = "账单"
    #
    new_directory = createDirectory(current_directory,output_folder_name)
    bill_file= merge_bill(bill_folder_name)
    # bill_file = os.path.join(new_directory, str(current_year % 100) + "年" + str(current_month) + "月账单核算.xlsx")
    BillPivotTable_file = pivot(bill_file, current_directory, output_folder_name)
    # BillPivotTable_file = os.path.join(current_directory, output_folder_name, "Merge_BillPivotTable.xlsx")
    merge_all(bill_file, BillPivotTable_file)

    message = "完成！"
    title = "账单核算报表完成"
    pyautogui.alert(message, title)
