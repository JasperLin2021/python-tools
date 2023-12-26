import os
import csv
from datetime import datetime

import pandas as pd
import pyautogui
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from utils import get_ad_sku_dict


def createDirectory(current_directory):
    # 新建文件夹
    folder_name = "输出"
    folder_path = os.path.join(current_directory, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        return folder_path
    return folder_path

def ad_handle(current_directory, new_directory, ad_folder_name):

    # 获取当前目录下的所有Excel文件
    files = [file for file in os.listdir(ad_folder_name) if file.endswith(
        '.csv') and not file.endswith(".png") and not file.startswith("E2")]

    df_account_list = []
    for file in files:
        file_path = os.path.join(current_directory, ad_folder_name, file)
        account = file.replace(" ", "").split("广告")[0].replace(".csv", "")

        df_account = pd.read_csv(file_path)

        df_account.insert(0, "账号", account)

        df_account["Item ID"] = df_account["Item ID"].astype(str)

        df_account_list.append(df_account)

    merged_df = pd.concat(df_account_list, ignore_index=True)
    xlsx_filename = os.path.join(new_directory, str(current_year % 100) + "年" + str(current_month) + "月广告核算.xlsx")
    merged_df.to_excel(xlsx_filename, index=False)


def bill_handle(current_directory, new_directory, bill_folder_name):
    files = [file for file in os.listdir(bill_folder_name) if file.endswith(
        '.csv')]

    df_account_list = []
    # 创建新的工作簿用于保存结果
    new_workbook = Workbook()
    new_worksheet = new_workbook.active

    merge_workbook = Workbook()
    merge_worksheet = merge_workbook.active

    for file in files:
        file_path = os.path.join(current_directory, bill_folder_name, file)
        account = file.replace(" ", "").split("账单")[0].replace(".csv", "")

        with open(file_path, encoding='utf-8-sig') as f:
            for i, row in enumerate(csv.reader(f, skipinitialspace=True), 1):
                if i >= 13:
                    df_account_list.append([account,row[17],'',row[0],'','',row[32].replace('-',''),'','',row[36]])
                    # print(row)

    df_account_list.insert(0, ['账号', '物品编号', '/', '交易创建日期', '/', '/', '交易总金额', '/', '/', '描述'])

    # 创建工作簿和工作表
    workbook = Workbook()
    sheet = workbook.active
    #
    # # 将列表数据写入工作表
    for row in df_account_list:
        sheet.append(row)
    #
    # # 保存工作簿为.xlsx文件
    # 获取列标题（第一行）的索引
    header_row = sheet[1]
    header_index = {header_row[i].value: i + 1 for i in range(len(header_row))}

    # 获取"描述"列的索引
    description_column_index = header_index.get("描述")

    # 遍历每一行，筛选出符合条件的内容
    for row in sheet.iter_rows(min_row=2):
        if row[description_column_index - 1].value == "Ad Fee Advanced ":
            # print(123)
            # new_worksheet.append([cell.value for cell in row])
            new_worksheet.append([row[0].value, row[1].value, '', row[3].value, '', '', float(row[6].value), '', ''])

    wb = load_workbook(
        os.path.join(new_directory, str(current_year % 100) + "年" + str(current_month) + "月广告核算.xlsx"))
    ws = wb.active

    for row in ws.iter_rows():
        merge_row = [cell.value for cell in row]
        merge_worksheet.append(merge_row)

    for row in new_worksheet.iter_rows():
        merge_row = [cell.value for cell in row]
        merge_worksheet.append(merge_row)


    #添加运营人员
    merge_worksheet['I1'] = "运营"
    id_operator_file = "ID_运营表.xlsx"
    id_operator_dict = get_ad_sku_dict(id_operator_file, "已删除重复项", 0, 2)
    # 获取数据行数
    row_count = merge_worksheet.max_row
    # 从I2单元格开始，逐行填入递增的数字
    for row in range(2, row_count + 1):
        cell = f'I{row}'
        # print(cell)
        merge_worksheet[cell].value = id_operator_dict.get(merge_worksheet[f'B{row}'].value, '-')[0]

    merge_worksheet.title = "总表"
    merge_workbook.save(os.path.join(new_directory, str(current_year % 100) + "年" + str(current_month) + "月广告核算.xlsx"))


if __name__ == '__main__':
    # 获取当前日期和时间
    current_date = datetime.now()
    # 获取当前年份
    current_year = current_date.year
    # 获取当前月份
    current_month = current_date.month
    current_directory = os.getcwd()

    output_folder_name = "输出"
    ad_folder_name = "广告"
    bill_folder_name = "账单"

    new_directory = createDirectory(current_directory)

    ad_handle(current_directory, new_directory, ad_folder_name)
    bill_handle(current_directory, new_directory, bill_folder_name)

    message = "完成！"
    title = "广告核算报表完成"
    pyautogui.alert(message, title)