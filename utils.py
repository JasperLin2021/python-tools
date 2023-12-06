from datetime import datetime

import openpyxl
from openpyxl.reader.excel import load_workbook


def deleteRow(file, type, keyword=None, column=None):
    wb = load_workbook(file)

    # 删除空白行
    if type == 1:
        # 遍历所有工作表
        for ws in wb.worksheets:
            rows = list(ws.rows)
            for row in reversed(rows):
                if row[0].value is None:
                    ws.delete_rows(row[0].row)
    # 删除第一行
    elif type == 2:
        # 遍历所有工作表
        for ws in wb.worksheets:
            ws.delete_rows(1)

    # 删除第二行及之后的包含"店铺"的行
    elif type == 3:
        for ws in wb.worksheets:
            rows = list(ws.rows)
            for row in rows[1:]:
                if row[0].value is not None and keyword in row[0].value:
                    ws.delete_rows(row[0].row)

    #从最后一行开始遍历每一行，如果某行的某列单元格的值不为空且包含指定关键字（keyword），则删除该行。
    elif type == 4:
        for ws in wb.worksheets:
            for row in reversed(range(1, ws.max_row + 1)):
                if ws.cell(row=row, column=column).value is not None and keyword in ws.cell(row=row, column=column).value:
                    ws.delete_rows(row)
    # 保存修改
    wb.save(file)


def unmergeCells(file):
    # 加载Excel文件
    workbook = load_workbook(file)

    # 遍历所有工作簿
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 获取所有合并单元格的范围和值
        merged_cells = list(sheet.merged_cells)  # 创建副本
        merged_values = {}

        # 遍历合并单元格
        for merged_range in merged_cells:
            # 获取合并单元格的范围边界
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

            # 获取合并单元格的值
            merged_value = sheet.cell(row=min_row, column=min_col).value

            # 保存合并单元格的值和范围
            merged_values[merged_range.coord] = {
                'value': merged_value,
                'range': (min_row, min_col, max_row, max_col)
            }

            # 取消合并单元格
            sheet.unmerge_cells(str(merged_range))

        # 恢复取消合并的单元格的值
        for coord, data in merged_values.items():
            merged_value = data['value']
            min_row, min_col, max_row, max_col = data['range']
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = merged_value

    # 保存修改后的Excel文件
    workbook.save(file)


def deleteCells(file, keyword):
    # 加载Excel文件
    wb = load_workbook(file)

    # 遍历所有工作表
    for ws in wb.worksheets:
        # 遍历每个单元格
        for row in ws.iter_rows():
            for cell in row:
                # 检查单元格内容是否包含"NewColumn"
                if cell.value and "NewColumn" in str(cell.value):
                    # 将内容设置为空
                    cell.value = None
                    cell.font = None
                    cell.border = None
                    cell.protection = None
                    cell.alignment = None

    # 保存修改
    wb.save(file)


def get_ad_sku_dict(file, sheetname, keycolum, valuecolum):
    ad_sku_dict = {}

    # 打开品牌广告明细sku.xlsx文件
    wb = openpyxl.load_workbook(file)
    # ws = wb.active

    for sheet_name in wb.sheetnames:
        if sheet_name == sheetname:
            ws = wb[sheet_name]

            # 遍历每行数据，构建字典
            for row in ws.iter_rows(min_row=2, values_only=True):
                ad_name = row[keycolum]
                ad_sku = row[valuecolum]

                if ad_name not in ad_sku_dict:
                    ad_sku_dict[ad_name] = [ad_sku]
                else:
                    ad_sku_dict[ad_name].append(ad_sku)

            return ad_sku_dict

def get_sku_site_fee_dict(file, sheetname, skucolum, sitecolum, fbadeliveryfeecolum):
    sku_site_fee_dict = {}

    # 打开品牌广告明细sku.xlsx文件
    wb = openpyxl.load_workbook(file)
    # ws = wb.active

    for sheet_name in wb.sheetnames:
        if sheet_name == sheetname:
            ws = wb[sheet_name]

            # 遍历每行数据，构建字典
            for row in ws.iter_rows(min_row=2, values_only=True):
                ad_sku = row[skucolum]
                ad_site= row[sitecolum]
                ad_fba_delivery_fee = row[fbadeliveryfeecolum]

                ad_site_dict = {}
                ad_site_dict[ad_site] = ad_fba_delivery_fee
                if ad_sku not in sku_site_fee_dict:
                    sku_site_fee_dict[ad_sku] = ad_site_dict
                else:
                    value = sku_site_fee_dict.get(ad_sku, {})
                    value.update(ad_site_dict)
                    sku_site_fee_dict[ad_sku] = value


            return sku_site_fee_dict

def getNumberDaysBetweenDates(sheetname, keyword):
    # 以"销售"作为分隔符，获取销售以后的所有字符
    sales_info = sheetname.split(keyword)[1]

    # 以"-"作为分隔符，获取两个日期字符串
    date_str1, date_str2 = sales_info.split("-")

    # 解析日期字符串为日期对象
    date_format = "%m.%d"
    date1 = datetime.strptime(date_str1, date_format).date()
    date2 = datetime.strptime(date_str2, date_format).date()

    # 计算日期差异
    date_diff = (date2 - date1).days

    return date_diff

def currencyConverter(cell):
    currency = get_ad_sku_dict("品牌广告明细sku.xlsx", "币种汇率", 0, 1)
    if cell.value is not None and cell.number_format.find("US$") != -1:
        result = float(str(cell.value)) * currency["US"][0]
        return result
    if cell.value is not None and str(cell.value).find("CA$") != -1:
        # print("字符串中包含'CA$'")
        result = float(str(cell.value).split("CA$")[1]) * currency["CA"][0]
        return result