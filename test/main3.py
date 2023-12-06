import os
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import Border, Side, Font, Alignment

from utils import deleteRow, get_ad_sku_dict
import pandas as pd


def filter():
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 新建文件夹
    folder_name2 = "输出2"
    folder_name3 = "输出3"

    # 创建一个新的Excel文件来保存数据透视表
    output_filename = "fba库存-new.xlsx"
    fba_raw_file = os.path.join(current_directory, folder_name2, output_filename)

    Startsellingdate_filename = "亚马逊产品开始出售日期.xlsx"
    Salesperson_dict = get_ad_sku_dict(Startsellingdate_filename, "Sheet1", 4, 4).keys()

    # 读取Excel文件
    df = pd.read_excel(fba_raw_file)

    for key in Salesperson_dict:
        # 根据条件筛选
        filtered_df = df[df.iloc[:, 0] == key]

        if filtered_df.size == 0:
            continue
        salesperson_file = os.path.join(current_directory, folder_name3, key + "-fba库存.xlsx")

        filtered_df.to_excel(salesperson_file, index=False)

        deleteRow(salesperson_file, 2)

        # 打开源文件
        source_file = os.path.join(current_directory, folder_name3, salesperson_file)
        source_workbook = openpyxl.load_workbook(source_file)
        source_sheet = source_workbook.active

        # 打开目标文件
        target_file = "模板-亚马逊库存分析.xlsx"
        target_workbook = openpyxl.load_workbook(target_file)
        target_sheet = target_workbook.active

        # 复制数据
        start_row = 4
        start_column = 1

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        for row in source_sheet.iter_rows():
            for cell in row:
                target_sheet.cell(row=start_row, column=start_column).value = cell.value
                target_sheet.cell(row=start_row, column=start_column).font = Font(name='宋体', size=11)

                target_sheet.cell(row=start_row, column=start_column).alignment = Alignment(vertical='center',
                                                                                            horizontal='center')
                target_sheet.cell(row=start_row, column=start_column).border = border
                if start_column in [29, 30, 31, 32, 33, 36, 37]:
                    target_sheet.cell(row=start_row, column=start_column).number_format = '0.00%'

                start_column += 1
            start_row += 1
            start_column = 1  # 重置列号B对应索引2

        rule1 = IconSetRule('3Arrows', "num", [-1, 0, 1], showValue=True, reverse=False)
        rule2 = IconSetRule('3Arrows', "num", [-0.1, 0, 0.1], showValue=True, reverse=False)

        target_sheet.conditional_formatting.add(f'S4:S{target_sheet.max_row}', rule1)
        target_sheet.conditional_formatting.add(f'W4:W{target_sheet.max_row}', rule2)
        target_sheet.conditional_formatting.add(f'AE4:AE{target_sheet.max_row}', rule2)

        # 冻结前3行
        target_sheet.freeze_panes = 'A4'

        # 设置第一行的行高为 150 磅
        target_sheet.row_dimensions[1].height = 150

        # 保存目标文件
        target_workbook.save(salesperson_file)


if __name__ == '__main__':
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    filter()

    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))